import requests
import pandas as pd
import openpyxl
from io import BytesIO

# Enlace de compartir de OneDrive (asegúrate de que permita descarga)
onedrive_url = "https://tauusbmededu-my.sharepoint.com/:x:/g/personal/jeferson_vega231_tau_usbmed_edu_co/IQDW2LL9iemtSZK_dociXutZAd8K4Khzle5pZY28X6-WSJU?e=1CsekW"

# Descargar el Excel desde OneDrive
response = requests.get(onedrive_url)
if response.status_code != 200:
    print("Error al descargar el archivo:", response.status_code)
    exit()

# Cargar el Excel en memoria
excel_data = BytesIO(response.content)
workbook = openpyxl.load_workbook(excel_data)

# Leer todas las hojas
all_data = []
for sheet_name in workbook.sheetnames:
    sheet = workbook[sheet_name]
    df = pd.read_excel(excel_data, sheet_name=sheet_name)
    # Normalizar columnas (similar a la lógica de JavaScript)
    df.columns = df.columns.str.strip().str.upper()
    df.rename(columns={
        'ID ESTUDIANTE': 'ID',
        'NOMBRE': 'Nombre',
        'PROGRAMA ACADÉMICO': 'Programa',
        'ULTIMA UBICACIÓN SEMESTRAL': 'Semestre'
    }, inplace=True)
    all_data.extend(df.to_dict('records'))

# Función para normalizar programa
def norm_prog(p):
    u = str(p or '').upper().replace('Á', 'A').replace('É', 'E').replace('Í', 'I').replace('Ó', 'O').replace('Ú', 'U').replace('Ñ', 'N')
    if 'ADMINISTR' in u:
        return 'ADM'
    elif 'CONTAD' in u:
        return 'CON'
    elif 'NEGOCIOS' in u:
        return 'NEG'
    return 'ADM'

# Asignar profesores equitativamente
profesores_por_carrera = {
    'NEG': ['María Victoria Botero', 'Wilson Ortegón'],
    'ADM': ['Andrés Arcos', 'Janeth Restrepo'],
    'CON': ['Sandra Patricia Arango']
}
all_profs = ['Andrés Arcos', 'Janeth Restrepo', 'María Victoria Botero', 'Wilson Ortegón', 'Sandra Patricia Arango']

def choose_prof(candidates, counts):
    min_count = min(counts[p] for p in candidates)
    for p in candidates:
        if counts[p] == min_count:
            return p

counts = {p: 0 for p in all_profs}
for estudiante in all_data:
    carrera = norm_prog(estudiante.get('Programa', ''))
    candidatos = profesores_por_carrera.get(carrera, [])
    if not candidatos:
        estudiante['Profesor'] = 'Sin asignar'
        continue

    min_global = min(counts[p] for p in all_profs)
    if carrera == 'CON':
        profesor_preferido = choose_prof(candidatos, counts)
        if counts[profesor_preferido] <= min_global + 1:
            estudiante['Profesor'] = profesor_preferido
        else:
            estudiante['Profesor'] = choose_prof(all_profs, counts)
    else:
        min_preferido = min(counts[p] for p in candidatos)
        if min_preferido > min_global + 1:
            estudiante['Profesor'] = choose_prof(all_profs, counts)
        else:
            estudiante['Profesor'] = choose_prof(candidatos, counts)

    counts[estudiante['Profesor']] += 1

# Crear nueva hoja de Asignaciones
asignaciones_data = [['Programa', 'ID Estudiante', 'Nombre', 'Profesor Asignado']]
for e in all_data:
    asignaciones_data.append([
        e.get('Programa', ''),
        e.get('ID', ''),
        e.get('Nombre', ''),
        e.get('Profesor', '')
    ])

# Agregar la nueva hoja al workbook
new_sheet = workbook.create_sheet('Asignaciones')
for row in asignaciones_data:
    new_sheet.append(row)

# Guardar el Excel actualizado localmente (el usuario lo sube manualmente a OneDrive)
output_path = 'Excel_Actualizado_Con_Asignaciones.xlsx'
workbook.save(output_path)
print(f"Excel actualizado guardado en: {output_path}")
print("Sube este archivo a OneDrive para reemplazar el original.")